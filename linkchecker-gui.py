"""
This program was written By Bret Wagner for the use of Drive Social Media
This was written to not have comments so that users were less likely to try to
change things that they don't need to change. At a later date I will add a
section near the top of the file where I will declare the variables you can
change in order to adjust the behavior of the script.

version 1.2.2

Changelog:
Version 1.2.2
* Added Padding around buttons to keep from clicking wrong one
Version 1.2.1:
* Fixed logic error that reset the empty cell counter on cells with non-link
data
Version 1.2:
* Fixed logic error to escape loop after 3 empty rows of values.

Version 1.1:
* Added error messages to remind user which modules still need installed

Version 1.0:
* First official release into office
"""

modules = []

try:
    from Tkinter import Tk
except:
    modules.append("Tkinter")
try:
    import ttk
except:
    modules.append("ttk")
try:
    import webbrowser
except:
    modules.append("webbrowser")
try:
    import openpyxl
except:
    modules.append("openpyxl")
try:
    import tkFileDialog
except:
    modules.append("tkFileDialog")


if modules:
    print("Things didn't install correctly " + str(modules))
    for mod in modules:
        print(mod)
        quit()


class MainWindow(ttk.Frame):
    def __init__(self, parent):
        ttk.Frame.__init__(self, parent)
        self.parent = parent
        labelspacer = ttk.Label(parent, width=30, text=" ")
        labelspacer.grid(column=1, row=3, sticky="W,E")
        browse_button = ttk.Button(parent, text="Browse...", command=browse)
        browse_button.grid(column=1, row=4, sticky="S")
        self.file = file
        label = ttk.Label(parent, text="""
        Click the browse button below
        to select your link spreadsheet         """)
        label.grid(column=1, row=2, sticky="S,W")
        label2 = ttk.Label(parent, text=" ")
        label2.grid(column=1, row=1, sticky="W,E")
        empty_space = ttk.Label(parent, text=" ")
        empty_space.grid(column=1, row=5)


class ClientList(ttk.Frame):
    def __init__(self, parent, clients):
        ttk.Frame.__init__(self, parent)
        self.parent = parent
        self.clients = clients[0]
        file = clients[1]
        row_count = 1
        column_count = 1
        for client in self.clients:
            client_button = ttk.Button(parent, text=client,
                        command=lambda client=client: openLink(client, file))
            client_button.grid(column=column_count,
                               row=row_count,
                               sticky="W,E",
                               padx=5,
                               pady=5)
            column_count += 1
            if column_count % 4 == 0:
                row_count += 1
                column_count = 1


def openLink(client, file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(client)
    count = 0
    empty_row = 0
    for row in sheet.iter_rows('B2:Z40'): # change the range of rows and columns here
        empty_cell = 0
        if empty_row <= 2:
            for cell in row:
                if str(cell.value).startswith('http'):
                    webbrowser.open(cell.value, autoraise=False)
                    count += 1
                    empty_cell = 0
                elif not str(cell.value).startswith('http'):
                    empty_cell += 1
                    if empty_cell >= 25 and empty_row <= 2:
                        empty_row += 1
                else:
                    break
        elif empty_row > 2:
            break

    print("Thanks")
    print("Opened " + str(count) + " Links")


def browse():
    file = tkFileDialog.askopenfilename()
    clients = check(file)
    client_list = Tk()
    root.destroy()
    client_list.title("Client List")
    ClientList(client_list, clients=clients)
    client_list.mainloop()


def check(file):
    clients = []
    workbook = openpyxl.load_workbook(file)
    sheets = workbook.get_sheet_names()
    for client in sheets:
        clients.append(client)

    return clients, file

if __name__ == '__main__':
    root = Tk()
    root.title("Choose a file")
    MainWindow(root)
    root.mainloop()
